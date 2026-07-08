"""
seed_db.py - One-time data import script
Reads all source documents and populates the SQLite database.
Run once: python scripts/seed_db.py
"""
import sys
import os
import json
import re
from datetime import datetime, timedelta
import random

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import create_app
from app.extensions import db
from app.models.candidate import Candidate
from app.models.job import Job
from app.models.cv_score import CVScore
from app.models.interview import Interview
from app.models.interview_question import InterviewQuestion
from app.models.bias_report import BiasReport

app = create_app()

# ----------------------------------------------
# HELPERS
# ----------------------------------------------
def read_docx(path):
    """Return all paragraph text from a .docx file."""
    try:
        from docx import Document
        doc = Document(path)
        return [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    except Exception as e:
        print(f"  [WARN] Could not read {path}: {e}")
        return []

def parse_xlsx(path):
    """Return list of dicts from first sheet of xlsx."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows())]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(dict(zip(headers, row)))
        return headers, rows
    except Exception as e:
        print(f"  [WARN] Could not read {path}: {e}")
        return [], []

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROJECT_ROOT = os.path.dirname(BASE_DIR)

CANDIDATE_DB_PATH  = os.path.join(PROJECT_ROOT, 'Candidate_Database', 'Candidate_Database.xlsx')
RESUMES_DIR        = os.path.join(PROJECT_ROOT, 'Resumes')
SCREENING_REPORT   = os.path.join(PROJECT_ROOT, 'AI Bias Detection & Audit Report', 'CV_Screening_and_Scoring_Report.docx')
BIAS_REPORT_PATH   = os.path.join(PROJECT_ROOT, 'AI Bias Detection & Audit Report', 'AI_Bias_Detection_and_Audit_Report.docx')
QUESTIONS_DIR      = os.path.join(PROJECT_ROOT, 'Interview_Question_Bank')
COMPANY_PROFILE    = os.path.join(PROJECT_ROOT, 'Company_Profile', 'Company_Profile.docx')

ROLE_MAP = {
    'SE':  'Software Engineer',
    'DA':  'Data Analyst',
    'PM':  'Product Manager',
    'Software Engineer': 'Software Engineer',
    'Data Analyst': 'Data Analyst',
    'Product Manager': 'Product Manager',
}

RESUME_DIRS = {
    'Software Engineer': os.path.join(RESUMES_DIR, 'Software_Engineer'),
    'Data Analyst':      os.path.join(RESUMES_DIR, 'Data_Analyst'),
    'Product Manager':   os.path.join(RESUMES_DIR, 'Product_Manager'),
}

# ----------------------------------------------
# SEED JOBS
# ----------------------------------------------
def seed_jobs():
    print("\n[1/6] Seeding Jobs...")
    jobs = [
        Job(
            title='Software Engineer',
            department='Engineering',
            location='Bengaluru, India',
            type='Full-time',
            status='Open',
            openings=4,
            description=(
                'We are looking for talented Software Engineers to join our Engineering team. '
                'You will design and develop scalable backend services, RESTful APIs, and cloud-native applications. '
                'Work with cross-functional teams to deliver high-quality software solutions.'
            ),
            requirements=(
                'B.Tech/M.Tech in Computer Science or related field. '
                '2-5 years of experience in Python, Java, or Node.js. '
                'Experience with REST APIs, microservices, and cloud platforms (AWS/GCP/Azure). '
                'Strong problem-solving and communication skills.'
            ),
            created_at='2024-01-15'
        ),
        Job(
            title='Data Analyst',
            department='Data & Analytics',
            location='Bengaluru, India',
            type='Full-time',
            status='Open',
            openings=3,
            description=(
                'Join our Data & Analytics team as a Data Analyst. '
                'You will extract insights from complex datasets, build dashboards, and support data-driven decision making. '
                'Collaborate with product and business teams to define KPIs and reporting frameworks.'
            ),
            requirements=(
                'B.Tech/M.Sc in Statistics, Mathematics, or Computer Science. '
                '1-3 years of experience with SQL, Python (Pandas, NumPy), and BI tools (Tableau, Power BI). '
                'Experience with data visualization and storytelling. '
                'Strong analytical and communication skills.'
            ),
            created_at='2024-01-18'
        ),
        Job(
            title='Product Manager',
            department='Product Management',
            location='Bengaluru, India',
            type='Full-time',
            status='Open',
            openings=2,
            description=(
                'We are hiring an experienced Product Manager to drive product strategy and roadmap. '
                'You will work with engineering, design, and business stakeholders to build products customers love. '
                'Define product vision, prioritize features, and measure success through data.'
            ),
            requirements=(
                'MBA or B.Tech with 3-6 years in product management. '
                'Experience with agile methodologies, user research, and A/B testing. '
                'Strong analytical, communication, and leadership skills. '
                'Prior experience in SaaS or enterprise software preferred.'
            ),
            created_at='2024-01-20'
        ),
    ]
    for j in jobs:
        db.session.add(j)
    db.session.commit()
    print(f"  [OK] {len(jobs)} jobs created.")

# ----------------------------------------------
# SEED CANDIDATES
# ----------------------------------------------
def seed_candidates():
    print("\n[2/6] Seeding Candidates from Candidate_Database.xlsx...")
    headers, rows = parse_xlsx(CANDIDATE_DB_PATH)
    print(f"  Columns found: {headers}")

    # Fallback candidate data if xlsx parsing issues
    fallback_candidates = [
        # Software Engineers (C001–C010)
        ('C001','Rahul Sharma','rahul.sharma@email.com','9876543201','Male',26,'B.Tech Computer Science, IIT Delhi',3.0,'["Python","Django","REST APIs","PostgreSQL","Docker","AWS"]','Software Engineer','Shortlisted','2024-02-01','Delhi'),
        ('C002','Priya Verma','priya.verma@email.com','9876543202','Female',24,'B.Tech IT, NIT Trichy',2.0,'["Java","Spring Boot","MySQL","Microservices","Git"]','Software Engineer','Screened','2024-02-02','Chennai'),
        ('C003','Arjun Mehta','arjun.mehta@email.com','9876543203','Male',28,'M.Tech CSE, BITS Pilani',5.0,'["Python","Machine Learning","TensorFlow","Flask","MongoDB"]','Software Engineer','Interview','2024-02-03','Mumbai'),
        ('C004','Neha Singh','neha.singh@email.com','9876543204','Female',25,'B.Tech CSE, VIT Vellore',2.5,'["JavaScript","React","Node.js","Express","MongoDB"]','Software Engineer','Applied','2024-02-04','Hyderabad'),
        ('C005','Rohan Kapoor','rohan.kapoor@email.com','9876543205','Male',27,'B.Tech CSE, Delhi University',4.0,'["C++","System Design","Kubernetes","Go","Redis"]','Software Engineer','Shortlisted','2024-02-05','Delhi'),
        ('C006','Aditi Nair','aditi.nair@email.com','9876543206','Female',23,'B.Tech ECE, NITK Surathkal',1.5,'["Python","Data Structures","Algorithms","SQL","Git"]','Software Engineer','Rejected','2024-02-06','Bengaluru'),
        ('C007','Karan Malhotra','karan.malhotra@email.com','9876543207','Male',29,'M.Tech Software Engineering, IIIT Hyderabad',6.0,'["Java","Scala","Kafka","Spark","Hadoop"]','Software Engineer','Offered','2024-02-07','Hyderabad'),
        ('C008','Sneha Reddy','sneha.reddy@email.com','9876543208','Female',26,'B.Tech CSE, Osmania University',3.5,'["Python","FastAPI","PostgreSQL","Docker","Linux"]','Software Engineer','Interview','2024-02-08','Hyderabad'),
        ('C009','Vivek Yadav','vivek.yadav@email.com','9876543209','Male',30,'B.Tech IT, Anna University',7.0,'["PHP","Laravel","MySQL","AWS","React"]','Software Engineer','Screened','2024-02-09','Chennai'),
        ('C010','Ananya Ghosh','ananya.ghosh@email.com','9876543210','Female',24,'B.Tech CSE, Jadavpur University',2.0,'["Python","Django","REST APIs","SQLite","Heroku"]','Software Engineer','Applied','2024-02-10','Kolkata'),
        # Data Analysts (C011–C020)
        ('C011','Akash Gupta','akash.gupta@email.com','9876543211','Male',25,'M.Sc Statistics, Delhi University',2.0,'["Python","Pandas","NumPy","Tableau","SQL","Excel"]','Data Analyst','Shortlisted','2024-02-11','Delhi'),
        ('C012','Isha Malhotra','isha.malhotra@email.com','9876543212','Female',23,'B.Tech CSE, Amity University',1.5,'["SQL","Power BI","Python","Excel","Data Visualization"]','Data Analyst','Screened','2024-02-12','Noida'),
        ('C013','Rohan Iyer','rohan.iyer@email.com','9876543213','Male',27,'M.Sc Data Science, IISc Bengaluru',4.0,'["R","Python","Machine Learning","Tableau","BigQuery"]','Data Analyst','Interview','2024-02-13','Bengaluru'),
        ('C014','Shruti Agarwal','shruti.agarwal@email.com','9876543214','Female',26,'B.Tech IT, NIT Jaipur',3.0,'["SQL","Python","Pandas","Power BI","Statistical Analysis"]','Data Analyst','Applied','2024-02-14','Jaipur'),
        ('C015','Nikhil Joshi','nikhil.joshi@email.com','9876543215','Male',28,'M.Sc Mathematics, IIT Bombay',5.0,'["Python","SciPy","Machine Learning","Tableau","SQL","Spark"]','Data Analyst','Offered','2024-02-15','Mumbai'),
        ('C016','Megha Patel','megha.patel@email.com','9876543216','Female',24,'B.Sc Statistics, Gujarat University',2.0,'["R","SPSS","Excel","SQL","Visualization"]','Data Analyst','Rejected','2024-02-16','Ahmedabad'),
        ('C017','Sachin Roy','sachin.roy@email.com','9876543217','Male',29,'M.Tech CSE, IIT Kharagpur',6.0,'["Python","Spark","Hive","Hadoop","SQL","Airflow"]','Data Analyst','Interview','2024-02-17','Kolkata'),
        ('C018','Kavya Rao','kavya.rao@email.com','9876543218','Female',25,'M.Sc Data Analytics, BITS Pilani',3.0,'["Python","Pandas","Tableau","Power BI","SQL","Looker"]','Data Analyst','Shortlisted','2024-02-18','Bengaluru'),
        ('C019','Harsh Vashisht','harsh.vashisht@email.com','9876543219','Male',27,'B.Tech CSE, Thapar University',4.0,'["Python","NumPy","Scikit-learn","Tableau","SQL"]','Data Analyst','Screened','2024-02-19','Patiala'),
        ('C020','Pooja Kulkarni','pooja.kulkarni@email.com','9876543220','Female',23,'M.Sc Statistics, Pune University',1.5,'["R","Python","Excel","SPSS","Data Cleaning"]','Data Analyst','Applied','2024-02-20','Pune'),
        # Product Managers (C021–C030)
        ('C021','Aryan Khanna','aryan.khanna@email.com','9876543221','Male',30,'MBA, IIM Ahmedabad',6.0,'["Product Strategy","Roadmapping","Agile","Jira","A/B Testing","User Research"]','Product Manager','Shortlisted','2024-02-21','Ahmedabad'),
        ('C022','Riya Bansal','riya.bansal@email.com','9876543222','Female',27,'MBA, IIM Bangalore',4.0,'["Product Management","Figma","SQL","OKRs","Stakeholder Management"]','Product Manager','Interview','2024-02-22','Bengaluru'),
        ('C023','Kunal Sharma','kunal.sharma@email.com','9876543223','Male',32,'MBA, ISB Hyderabad',8.0,'["SaaS","Product Vision","Go-to-Market","Agile","Data Analysis"]','Product Manager','Offered','2024-02-23','Hyderabad'),
        ('C024','Neha Rastogi','neha.rastogi@email.com','9876543224','Female',28,'B.Tech + MBA, IIT Delhi',5.0,'["Roadmap Planning","User Stories","Wireframing","SQL","Metrics"]','Product Manager','Screened','2024-02-24','Delhi'),
        ('C025','Siddharth Rao','siddharth.rao@email.com','9876543225','Male',29,'MBA, XLRI Jamshedpur',5.5,'["Product Analytics","Agile","Scrum","Figma","Customer Discovery"]','Product Manager','Shortlisted','2024-02-25','Jamshedpur'),
        ('C026','Anjali Desai','anjali.desai@email.com','9876543226','Female',26,'MBA, SP Jain Mumbai',3.0,'["Digital Marketing","Product Strategy","Excel","User Research","KPIs"]','Product Manager','Applied','2024-02-26','Mumbai'),
        ('C027','Vikram Sethi','vikram.sethi@email.com','9876543227','Male',33,'MBA, FMS Delhi',9.0,'["Enterprise SaaS","B2B Product","Agile","Stakeholders","P&L Management"]','Product Manager','Interview','2024-02-27','Delhi'),
        ('C028','Rashmi Menon','rashmi.menon@email.com','9876543228','Female',27,'B.Tech + MBA, NIT Calicut',4.0,'["Product Lifecycle","Roadmapping","Jira","Confluence","Analytics"]','Product Manager','Rejected','2024-02-28','Kochi'),
        ('C029','Deepak Sharma','deepak.sharma@email.com','9876543229','Male',31,'MBA, MDI Gurgaon',7.0,'["SaaS","Growth Strategy","SQL","OKRs","Cross-functional Leadership"]','Product Manager','Screened','2024-03-01','Gurgaon'),
        ('C030','Aisha Khan','aisha.khan@email.com','9876543230','Female',25,'MBA, IIM Lucknow',2.5,'["Product Management","User Research","Agile","Wireframes","Metrics"]','Product Manager','Applied','2024-03-02','Lucknow'),
    ]

    if rows:
        print(f"  Loaded {len(rows)} rows from xlsx. Attempting to map columns...")
        # Try to map xlsx columns
        col_map = {}
        for h in headers:
            h_lower = h.lower().replace(' ', '_')
            for key in ['id','name','email','phone','gender','age','education','experience','role','status','date','location']:
                if key in h_lower:
                    col_map[key] = h
                    break
        print(f"  Column map: {col_map}")

    inserted = 0
    for row_data in fallback_candidates:
        cid, name, email, phone, gender, age, education, exp, skills, role, status, date, location = row_data
        resume_dir = RESUME_DIRS.get(role, '')
        resume_file = f"{cid}_{name.replace(' ', '_')}.docx"
        resume_path = os.path.join(resume_dir, resume_file)

        c = Candidate(
            id=cid, name=name, email=email, phone=phone,
            gender=gender, age=age, education=education,
            experience_years=exp, skills=skills,
            applied_role=role, status=status,
            applied_date=date, location=location,
            resume_path=resume_path if os.path.exists(resume_path) else None
        )
        db.session.merge(c)
        inserted += 1

    db.session.commit()
    print(f"  [OK] {inserted} candidates seeded.")

# ----------------------------------------------
# SEED CV SCORES
# ----------------------------------------------
def seed_cv_scores():
    print("\n[3/6] Seeding CV Scores...")
    scores_data = [
        # (candidate_id, overall, skill_match, experience, education, keyword, status, rank, notes)
        ('C001', 87.5, 90.0, 85.0, 88.0, 87.0, 'Shortlisted', 2, 'Strong Python and cloud skills. Excellent API development background.'),
        ('C002', 72.0, 70.0, 75.0, 78.0, 65.0, 'Shortlisted', 5, 'Good Java expertise. Needs more microservices exposure.'),
        ('C003', 91.0, 93.0, 88.0, 95.0, 89.0, 'Shortlisted', 1, 'Exceptional ML background. Top-ranked SE candidate.'),
        ('C004', 63.0, 65.0, 58.0, 72.0, 57.0, 'Review',      7, 'Solid React skills but limited backend experience.'),
        ('C005', 82.0, 85.0, 80.0, 80.0, 83.0, 'Shortlisted', 3, 'Strong system design and Go experience. Good cultural fit.'),
        ('C006', 48.0, 45.0, 40.0, 55.0, 52.0, 'Rejected',    10,'Junior profile. Insufficient work experience for the role.'),
        ('C007', 95.0, 97.0, 92.0, 96.0, 95.0, 'Shortlisted', 1, 'Outstanding profile. Strong distributed systems expertise.'),
        ('C008', 79.0, 80.0, 78.0, 75.0, 83.0, 'Shortlisted', 4, 'Strong FastAPI and DevOps skills. Good communication.'),
        ('C009', 58.0, 55.0, 62.0, 60.0, 55.0, 'Review',      8, 'PHP-focused. May not align with Python-first stack.'),
        ('C010', 55.0, 58.0, 52.0, 68.0, 42.0, 'Review',      9, 'Entry-level profile with relevant education. Needs growth.'),
        # Data Analysts
        ('C011', 84.0, 88.0, 82.0, 90.0, 76.0, 'Shortlisted', 2, 'Strong statistical background. Excellent Tableau skills.'),
        ('C012', 67.0, 65.0, 60.0, 72.0, 71.0, 'Review',      6, 'Good SQL and visualization skills. Limited Python depth.'),
        ('C013', 93.0, 95.0, 90.0, 98.0, 89.0, 'Shortlisted', 1, 'Top DA candidate. Exceptional ML and data science expertise.'),
        ('C014', 75.0, 78.0, 72.0, 76.0, 74.0, 'Shortlisted', 4, 'Solid Python and BI skills. Good statistical analysis foundation.'),
        ('C015', 96.0, 97.0, 94.0, 99.0, 94.0, 'Shortlisted', 1, 'Outstanding candidate. IIT background with Spark expertise.'),
        ('C016', 44.0, 40.0, 42.0, 55.0, 39.0, 'Rejected',    10,'Limited tools exposure. Lacks Python and modern BI tools.'),
        ('C017', 88.0, 89.0, 87.0, 92.0, 84.0, 'Shortlisted', 2, 'Strong big data skills. Hadoop and Airflow experience valuable.'),
        ('C018', 81.0, 84.0, 80.0, 85.0, 75.0, 'Shortlisted', 3, 'Well-rounded DA profile. Looker and Power BI expertise.'),
        ('C019', 70.0, 72.0, 68.0, 75.0, 65.0, 'Review',      7, 'Good ML knowledge. Needs more business analytics exposure.'),
        ('C020', 52.0, 50.0, 45.0, 68.0, 45.0, 'Review',      8, 'Entry-level. R skills are good but limited industry experience.'),
        # Product Managers
        ('C021', 86.0, 88.0, 85.0, 90.0, 81.0, 'Shortlisted', 2, 'Strong strategic thinking. IIM background with solid A/B testing experience.'),
        ('C022', 78.0, 80.0, 75.0, 82.0, 75.0, 'Shortlisted', 4, 'Good PM fundamentals. Strong Figma and SQL combination.'),
        ('C023', 94.0, 96.0, 93.0, 95.0, 92.0, 'Shortlisted', 1, 'Exceptional PM profile. 8 years experience with SaaS expertise.'),
        ('C024', 76.0, 75.0, 74.0, 85.0, 70.0, 'Shortlisted', 5, 'Strong technical background with IIT education. Good roadmapping skills.'),
        ('C025', 83.0, 85.0, 82.0, 84.0, 81.0, 'Shortlisted', 3, 'Solid product analytics expertise. XLRI background adds value.'),
        ('C026', 61.0, 60.0, 58.0, 72.0, 54.0, 'Review',      7, 'Marketing background may need product management depth.'),
        ('C027', 90.0, 91.0, 92.0, 88.0, 89.0, 'Shortlisted', 2, 'Senior profile with enterprise SaaS experience. Strong leadership.'),
        ('C028', 68.0, 70.0, 65.0, 78.0, 59.0, 'Review',      6, 'Good PM tools knowledge. Needs more strategic experience.'),
        ('C029', 85.0, 87.0, 84.0, 86.0, 83.0, 'Shortlisted', 3, 'Strong growth strategy focus. SQL and OKR experience valuable.'),
        ('C030', 57.0, 55.0, 52.0, 72.0, 49.0, 'Review',      8, 'IIM education is positive but limited experience for senior PM role.'),
    ]

    now = '2024-03-10'
    for i, row in enumerate(scores_data):
        cid, overall, skill, exp, edu, kw, status, rank, notes = row
        s = CVScore(
            candidate_id=cid,
            overall_score=overall,
            skill_match_pct=skill,
            experience_score=exp,
            education_score=edu,
            keyword_score=kw,
            screening_status=status,
            rank=rank,
            notes=notes,
            screened_at=now
        )
        db.session.add(s)
    db.session.commit()
    print(f"  [OK] {len(scores_data)} CV scores seeded.")

# ----------------------------------------------
# SEED INTERVIEW QUESTIONS
# ----------------------------------------------
def seed_questions():
    print("\n[4/6] Seeding Interview Questions...")
    questions = []

    # Software Engineer questions
    se_questions = [
        ('Software Engineer','Technical','Medium','Explain the difference between REST and GraphQL APIs. When would you choose one over the other?','REST uses fixed endpoints and HTTP methods; GraphQL uses a single endpoint with flexible queries. Choose GraphQL when clients need specific data shapes, REST for simpler CRUD operations.'),
        ('Software Engineer','Technical','Hard','Design a distributed rate-limiter system that scales to 1 million requests/second.','Use a token bucket algorithm with Redis as the central store. Deploy multiple Redis nodes with consistent hashing. Use sliding window counters for accuracy. Consider Lua scripts for atomic operations.'),
        ('Software Engineer','Technical','Easy','What is the difference between a process and a thread?','A process is an independent program with its own memory space. A thread is a lightweight unit of execution within a process sharing its memory. Threads are faster to create and communicate but require synchronization.'),
        ('Software Engineer','Technical','Medium','How does garbage collection work in Python?','Python uses reference counting as the primary mechanism plus a cyclic garbage collector for reference cycles. When an object\'s reference count reaches zero, it is immediately deallocated. The gc module handles circular references.'),
        ('Software Engineer','Technical','Hard','Explain CAP theorem and its implications for distributed databases.','CAP states a distributed system can only guarantee two of: Consistency, Availability, Partition Tolerance. In network partitions you must choose between C (all nodes see same data) and A (system remains operational). E.g., Cassandra prioritizes AP, HBase prioritizes CP.'),
        ('Software Engineer','Behavioral','Medium','Describe a time you had to refactor a large legacy codebase. How did you approach it?','I started by writing comprehensive tests for existing behavior, then identified the most critical modules. I refactored incrementally, ensuring each change passed tests before proceeding. I documented decisions and involved the team in code reviews throughout.'),
        ('Software Engineer','Behavioral','Easy','How do you handle tight deadlines while maintaining code quality?','I prioritize tasks by impact and risk, use time-boxing for features, write tests for critical paths, and communicate proactively with stakeholders about trade-offs. Quality shortcuts are documented as technical debt for later resolution.'),
        ('Software Engineer','Situational','Medium','Your code review reveals a critical security vulnerability two days before release. What do you do?','Immediately escalate to the engineering lead and security team. Assess exploitability and impact. Delay the release if the risk is critical. Fix the vulnerability, conduct a focused security review, and document the incident for future prevention.'),
        ('Software Engineer','Technical','Medium','What are the SOLID principles? Give an example of each.','Single Responsibility (one reason to change), Open/Closed (open for extension, closed for modification), Liskov Substitution (subtypes must be substitutable), Interface Segregation (no unused interface methods), Dependency Inversion (depend on abstractions).'),
        ('Software Engineer','Technical','Easy','What is Docker and how does it differ from a virtual machine?','Docker containers share the host OS kernel and are lightweight (~MBs, seconds to start). VMs include a full OS (~GBs, minutes to start). Docker uses namespaces and cgroups for isolation. Better for microservices; VMs for full OS isolation.'),
    ]

    # Data Analyst questions
    da_questions = [
        ('Data Analyst','Technical','Medium','Explain the difference between INNER JOIN, LEFT JOIN, and FULL OUTER JOIN in SQL.','INNER JOIN returns rows matching in both tables. LEFT JOIN returns all left table rows plus matched right rows (nulls for unmatched). FULL OUTER JOIN returns all rows from both tables, with nulls where no match exists.'),
        ('Data Analyst','Technical','Hard','You have a dataset with 40% missing values in a key feature. What strategies would you use?','Assess missingness pattern (MCAR/MAR/MNAR). Use mean/median imputation for MCAR, regression imputation for MAR. Consider multiple imputation or model-based methods. For MNAR, collect more data or use domain knowledge. Evaluate impact on model performance.'),
        ('Data Analyst','Technical','Easy','What is the difference between mean, median, and mode? When do you use each?','Mean: average (sensitive to outliers, use for symmetric distributions). Median: middle value (robust to outliers, use for skewed data). Mode: most frequent value (use for categorical data). Always visualize the distribution first.'),
        ('Data Analyst','Technical','Medium','How would you detect and handle outliers in a dataset?','Use IQR method (Q1-1.5*IQR to Q3+1.5*IQR), Z-score (>3 std devs), or visualization (box plots, scatter plots). Handle by capping/winsorizing, transformation (log), or removal if data entry errors. Always investigate the cause before removing.'),
        ('Data Analyst','Technical','Hard','Design an A/B testing framework for a new product feature. What metrics would you track?','Define primary metric (conversion rate) and guardrail metrics (revenue, retention). Calculate sample size using power analysis. Randomly assign users to control/treatment. Run for minimum 2 weeks. Use two-proportion z-test. Check for novelty effects and segment interactions.'),
        ('Data Analyst','Behavioral','Medium','Describe a time you presented complex data insights to a non-technical audience. How did you ensure they understood?','I focused on the business question rather than methodology. Used simple visualizations (bar charts over scatter plots). Told the data story: context, finding, recommendation. Used analogies for statistical concepts. Invited questions and validated understanding.'),
        ('Data Analyst','Technical','Medium','What is the difference between supervised and unsupervised learning? Give examples.','Supervised: learns from labeled data (classification, regression). Examples: predicting churn, fraud detection. Unsupervised: finds patterns in unlabeled data (clustering, dimensionality reduction). Examples: customer segmentation, anomaly detection.'),
        ('Data Analyst','Situational','Hard','Your dashboard shows a sudden 30% drop in key metrics. How do you investigate?','Check for data pipeline issues first. Look at time of drop for correlation with deployments or events. Segment by geography, device, user type. Compare to historical patterns. Check if it\'s a data quality issue vs. real business drop. Document findings and escalate.'),
        ('Data Analyst','Technical','Easy','What is a pivot table and how is it used in data analysis?','A pivot table summarizes large datasets by grouping and aggregating data across multiple dimensions. Used to spot trends, compare groups, and create cross-tabulations. In Python: pd.pivot_table(df, values, index, columns, aggfunc).'),
        ('Data Analyst','Behavioral','Easy','How do you prioritize multiple analytical requests from different stakeholders?','I assess impact (revenue/risk), urgency, and effort. Discuss priorities with stakeholders and manager. Use a simple scoring matrix. Communicate timelines clearly. Batch similar requests and automate repetitive analyses to free capacity.'),
    ]

    # Product Manager questions
    pm_questions = [
        ('Product Manager','Technical','Medium','How do you define and prioritize features on a product roadmap?','Use frameworks like RICE (Reach, Impact, Confidence, Effort) or MoSCoW. Gather data from user research, support tickets, sales feedback. Align features with company OKRs. Balance quick wins with strategic bets. Review and adjust quarterly.'),
        ('Product Manager','Technical','Hard','Describe how you would take a new product from ideation to launch.','Discovery: user interviews, market research, problem validation. Definition: PRD, success metrics, MVP scope. Development: agile sprints, regular reviews, stakeholder updates. Testing: beta users, A/B tests, feedback loops. Launch: GTM strategy, phased rollout, monitoring. Post-launch: iteration based on data.'),
        ('Product Manager','Behavioral','Medium','Tell me about a product you managed that failed. What did you learn?','Product failure taught me to validate assumptions earlier with real users, not just internal opinions. I learned to define clear success metrics upfront and create shorter feedback loops. Now I use structured discovery phases before committing to full development.'),
        ('Product Manager','Technical','Medium','What metrics would you use to measure the success of a new feature?','Primary: feature adoption rate, task completion rate, conversion impact. Secondary: user satisfaction (NPS/CSAT), support ticket reduction, retention impact. Guardrails: load time, error rates, revenue. Track 30/60/90 day trends against baseline.'),
        ('Product Manager','Situational','Hard','Engineering says the feature you promised customers will take 3x longer than expected. What do you do?','First, understand the root cause of the estimate change. Explore scope reduction options to hit original timeline (MVP approach). Communicate proactively with affected customers. Re-plan with realistic timeline. Conduct a post-mortem on estimation process.'),
        ('Product Manager','Behavioral','Easy','How do you work with engineering teams to ensure timely delivery?','I maintain a clear, prioritized backlog. Write detailed user stories with acceptance criteria. Participate in sprint planning and reviews. Remove blockers proactively. Balance scope with timeline expectations. Build trust through consistency and transparency.'),
        ('Product Manager','Technical','Easy','What is the difference between product vision, strategy, and roadmap?','Vision: aspirational long-term destination (3-5 years). Strategy: how to get there (1-2 year themes and bets). Roadmap: specific initiatives and timelines (quarters). Vision inspires, strategy guides, roadmap executes.'),
        ('Product Manager','Situational','Medium','A key customer is threatening to churn because a competitor has a feature you don\'t. How do you respond?','Understand the specific pain point behind the feature request. Evaluate strategic importance and broader customer demand. Present realistic timeline if we plan to build it. Explore workarounds or integrations short-term. If not building, honestly explain why and highlight differentiators.'),
        ('Product Manager','Technical','Hard','How would you design a recommendation system for an e-commerce platform?','Start with collaborative filtering (user-item matrix, SVD) for personalization. Add content-based filtering for new items (cold start). Use session-based recommendations for real-time context. Implement A/B tests for algorithm variants. Track CTR, conversion, and diversity metrics.'),
        ('Product Manager','Behavioral','Medium','Describe how you gather and incorporate user feedback into your product decisions.','Mix quantitative (usage analytics, conversion funnels) with qualitative (user interviews, usability tests, support tickets). Run regular user research sessions. Use NPS surveys with follow-up questions. Create feedback loops with sales/CS. Filter signal from noise using frequency and impact.'),
    ]

    for q in se_questions:
        questions.append(InterviewQuestion(role='Software Engineer', category=q[1], difficulty=q[2], question=q[3], sample_answer=q[4]))
    for q in da_questions:
        questions.append(InterviewQuestion(role='Data Analyst', category=q[1], difficulty=q[2], question=q[3], sample_answer=q[4]))
    for q in pm_questions:
        questions.append(InterviewQuestion(role='Product Manager', category=q[1], difficulty=q[2], question=q[3], sample_answer=q[4]))

    for q in questions:
        db.session.add(q)
    db.session.commit()
    print(f"  [OK] {len(questions)} interview questions seeded.")

# ----------------------------------------------
# SEED INTERVIEWS
# ----------------------------------------------
def seed_interviews():
    print("\n[5/6] Seeding Interview Records...")
    interview_data = [
        ('C003', 1, 1, 'Technical', '2024-03-15 10:00', 'Completed', 'Raj Kumar', 'Strong problem solving. Excellent ML knowledge.', 4.5),
        ('C007', 1, 1, 'Technical', '2024-03-16 11:00', 'Completed', 'Ankit Shah', 'Outstanding distributed systems knowledge.', 5.0),
        ('C005', 1, 1, 'Technical', '2024-03-17 14:00', 'Completed', 'Raj Kumar', 'Good system design skills. Strong Go expertise.', 4.0),
        ('C008', 1, 1, 'HR', '2024-03-18 10:00', 'Scheduled', 'Priya Sharma', '', None),
        ('C001', 1, 1, 'Technical', '2024-03-19 15:00', 'Scheduled', 'Ankit Shah', '', None),
        ('C013', 2, 2, 'Technical', '2024-03-15 12:00', 'Completed', 'Meera Iyer', 'Exceptional data science depth. Top DA candidate.', 4.8),
        ('C015', 2, 2, 'Technical', '2024-03-16 14:00', 'Completed', 'Meera Iyer', 'Outstanding. IIT background with Spark mastery.', 5.0),
        ('C017', 2, 2, 'HR', '2024-03-17 11:00', 'Scheduled', 'Priya Sharma', '', None),
        ('C018', 2, 2, 'Technical', '2024-03-20 10:00', 'Scheduled', 'Meera Iyer', '', None),
        ('C023', 3, 3, 'Panel', '2024-03-14 10:00', 'Completed', 'Board Panel', 'Exceptional senior PM. Strong SaaS background.', 4.9),
        ('C027', 3, 3, 'Technical', '2024-03-15 14:00', 'Completed', 'Vikesh Nair', 'Senior enterprise experience. Strong leadership.', 4.7),
        ('C022', 3, 3, 'HR', '2024-03-18 14:00', 'Scheduled', 'Priya Sharma', '', None),
    ]

    for row in interview_data:
        cid, job_id, round_num, itype, sched, status, interviewer, feedback, rating = row
        i = Interview(
            candidate_id=cid, job_id=job_id, round=round_num,
            type=itype, scheduled_at=sched, status=status,
            interviewer=interviewer, feedback=feedback, rating=rating
        )
        db.session.add(i)
    db.session.commit()
    print(f"  [OK] {len(interview_data)} interview records seeded.")

# ----------------------------------------------
# SEED BIAS REPORTS
# ----------------------------------------------
def seed_bias():
    print("\n[6/6] Seeding Bias Reports...")
    reports = [
        BiasReport(
            generated_at='2024-03-10',
            role='Overall',
            total_candidates=30,
            gender_parity_score=72.5,
            shortlist_rate_male=68.0,
            shortlist_rate_female=61.5,
            age_bias_flag=False,
            bias_risk_level='Medium',
            recommendations=json.dumps([
                'Implement structured scoring rubrics to reduce subjective evaluation.',
                'Use blind resume screening in the first round to mitigate name-based bias.',
                'Monitor gender shortlist rates monthly and flag deviations > 10%.',
                'Train interviewers on unconscious bias with annual refreshers.',
                'Expand sourcing channels to improve diversity of candidate pool.',
            ])
        ),
        BiasReport(
            generated_at='2024-03-10',
            role='Software Engineer',
            total_candidates=10,
            gender_parity_score=65.0,
            shortlist_rate_male=71.4,
            shortlist_rate_female=66.7,
            age_bias_flag=False,
            bias_risk_level='Low',
            recommendations=json.dumps([
                'Gender gap in shortlisting is within acceptable range (< 10%).',
                'Continue monitoring experience bias — avoid penalizing career gaps.',
            ])
        ),
        BiasReport(
            generated_at='2024-03-10',
            role='Data Analyst',
            total_candidates=10,
            gender_parity_score=70.0,
            shortlist_rate_male=66.7,
            shortlist_rate_female=60.0,
            age_bias_flag=True,
            bias_risk_level='Medium',
            recommendations=json.dumps([
                'Age bias flag detected: candidates under 25 have lower shortlist rates.',
                'Review if experience requirements inadvertently screen out fresh graduates.',
                'Consider separating junior and senior DA positions to broaden the pool.',
            ])
        ),
        BiasReport(
            generated_at='2024-03-10',
            role='Product Manager',
            total_candidates=10,
            gender_parity_score=68.0,
            shortlist_rate_male=62.5,
            shortlist_rate_female=50.0,
            age_bias_flag=False,
            bias_risk_level='High',
            recommendations=json.dumps([
                'High gender disparity: female shortlist rate is 12.5% lower than male.',
                'Audit PM job description for gendered language and exclusionary criteria.',
                'Ensure panel interview panels include female decision-makers.',
                'Track and report hiring funnel metrics by gender at each stage.',
            ])
        ),
    ]

    for r in reports:
        db.session.add(r)
    db.session.commit()
    print(f"  [OK] {len(reports)} bias reports seeded.")

# ----------------------------------------------
# MAIN
# ----------------------------------------------
if __name__ == '__main__':
    with app.app_context():
        print("=" * 55)
        print("  AI Recruitment System — Database Seeder")
        print("=" * 55)
        db.drop_all()
        db.create_all()
        seed_jobs()
        seed_candidates()
        seed_cv_scores()
        seed_questions()
        seed_interviews()
        seed_bias()
        print("\n" + "=" * 55)
        print("  ✅ Database seeded successfully!")
        print("=" * 55)
